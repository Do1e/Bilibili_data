# coding=gbk                        #1
import urllib.request               #2
import time                         #3
import openpyxl                     #4
import os                           #5
from name import get_name_of_video  #6
from bv_to_av import get_av         #7
from video_data import get_data     #8

'''
1 ת�����룬ʹ�����������������
2 ���ڶ�ȡ��ҳ��Ϣ
3 ���ڶ�ȡʱ��
4 ���ڶ�дExcel�ĵ�
5 �����ж��ĵ��Ƿ����
6 ��ȡ��Ƶ����
7 ��BV��תΪav��
8 ����ҳ��Ϣ��ȡ��Ƶ����
'''

# ��ȡ��Ƶʵʱ���ݲ�������Excel�ĵ���
def get_data_of_video():
    aid = bid = ''
    while(1):
        bid = input('������Ƶav�Ż�bv��:')
        if(bid[0:2] != 'av' and bid[0:2] != 'bv' and bid[0:2] != 'BV'):
            print('�������')
            continue
        name = get_name_of_video(bid)
        # ��ȡ��Ƶ����
        if(not name):
            print('�����ڸ���Ƶ')
        else:
            if(bid[0:2] != 'av'):
                aid = str(get_av(bid))
            else:
                aid = str(bid[2:])
            # ��ȡav��
            break
    name = get_name_of_video(bid)
    # ��ȡ��Ƶ����
    stoptime = input('����������ʱ��(��λ��):')
    stoptime = int(stoptime)
    # �������ʱ�䣬��λ����
    table_name = name + '.xlsx'
    # �����Excel�ļ���
    if(not os.path.exists(table_name)):
        File = openpyxl.Workbook()
        try:
            File.save(table_name)
        except OSError:
            print('���������ַ����޷���ԭ���Ʊ���')
            table_name = table_name.replace('\\', ' ')
            table_name = table_name.replace('/', ' ')
            table_name = table_name.replace(':', ' ')
            table_name = table_name.replace('*', ' ')
            table_name = table_name.replace('?', ' ')
            table_name = table_name.replace('"', ' ')
            table_name = table_name.replace('<', ' ')
            table_name = table_name.replace('>', ' ')
            table_name = table_name.replace('|', ' ')
            File.save(table_name)
    # ���������ļ������½��ļ�
    File = openpyxl.load_workbook(table_name)
    table = File[File.sheetnames[0]]
    # ��ȡ�ĵ��ĵ�һ�����
    if(not table['A1'].value):
        table['A1'] = 'ʱ��'
        table['B1'] = '����'
        table['C1'] = '����'
        table['D1'] = 'Ͷ��'
        table['E1'] = '�ղ�'
        table['F1'] = '��Ļ'
        table['G1'] = '����'
        table['H1'] = '����'
        table.column_dimensions['A'].width = 21
        Column = ['B', 'C', 'D', 'E', 'F', 'G', 'H']
        for i in Column:
            table.column_dimensions[i].width = 8
        # ���ñ���п�
        File.save(table_name)
    # ����ʱΪ�ձ������ɱ�ͷ�������п�
    url = 'http://api.bilibili.com/archive_stat/stat?aid=' + aid +'&type=jsonp'
    # ���ʸ���Ƶ���ݵ�api
    prettify = '{:20}\t{:9}\t{:5}\t{:5}\t{:5}\t{:5}\t{:5}\t{:5}'
    print(prettify.format('ʱ��', '����', '����', 'Ͷ��', '�ղ�', '��Ļ', '����', '����'))
    while(1):
        now_time_id = time.time()
        # ��ȡ��ǰʱ���
        now_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        # ����ʽ��ȡ��ǰʱ��
        page = urllib.request.urlopen(url)
        text = page.read()
        # ��ȡ��ҳ����
        text = text.decode('utf-8')
        view = get_data(text, 'view')
        like = get_data(text, 'like')
        coin = get_data(text, 'coin')
        favorite = get_data(text, 'favorite')
        danmaku = get_data(text, 'danmaku')
        share = get_data(text, 'share')
        reply = get_data(text, 'reply')
        # ��ȡ��������
        now_row = [now_time, view, like, coin, favorite, danmaku, share, reply]
        print(prettify.format(str(now_time), str(view), str(like), str(coin), str(favorite), str(danmaku), str(share), str(reply)))
        table.append(now_row)
        # ���ζ�ȡ������ĩβ
        File.save(table_name)
        while(1):
            if(time.time() - now_time_id >= stoptime):
                break
        # ��֤���δ��������ض�ʱ��
