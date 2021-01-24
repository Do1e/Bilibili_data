# coding=gbk                        #1
import urllib.request               #2
import time                         #3
import openpyxl                     #4
import os                           #5
from name import get_name_of_up     #6

'''
1 ת�����룬ʹ�����������������
2 ���ڶ�ȡ��ҳ��Ϣ
3 ���ڶ�ȡʱ��
4 ���ڶ�дExcel�ĵ�
5 �����ж��ĵ��Ƿ����
6 ��ȡ�ǳ�
'''

# ��ȡup��ʵʱ��˿����������Excel�ĵ���
def get_num_of_fans():
    while(1):
        uid = input('����Bվuid:')
        name = get_name_of_up(uid)
        if(not name):
            print('�����ڸ��û�')
        else:
            break
    # ��ȡ��up�����ǳ�
    stoptime = input('����������ʱ��(��λ��):')
    stoptime = int(stoptime)
    # �������ʱ�䣬��λ����
    table_name = name + '��ʵʱ��˿��.xlsx'
    # �����Excel�ļ���
    if(not os.path.exists(table_name)):
        File = openpyxl.Workbook()
        try:
            File.save(table_name)
        except OSError:
            print('���������ַ����޷���ԭ���Ʊ���')
            table_name = table_name.replace('\\', ' ')
            table_name = table_name.replace('/', ' ')
            table_name = table_name.replace(':', ' ')
            table_name = table_name.replace('*', ' ')
            table_name = table_name.replace('?', ' ')
            table_name = table_name.replace('"', ' ')
            table_name = table_name.replace('<', ' ')
            table_name = table_name.replace('>', ' ')
            table_name = table_name.replace('|', ' ')
            File.save(table_name)
    # ���������ļ������½��ļ�
    File = openpyxl.load_workbook(table_name)
    table = File[File.sheetnames[0]]
    # ��ȡ�ĵ��ĵ�һ�����
    if(not table['A1'].value):
        table['A1'] = 'ʱ��'
        table['B1'] = '��˿��'
        table.column_dimensions['A'].width = 21
        table.column_dimensions['B'].width = 8
        # ���ñ���п�
        File.save(table_name)
    # ����ʱΪ�ձ������ɱ�ͷ�������п�
    url = 'https://api.bilibili.com/x/relation/stat?vmid=' + uid +'&jsonp=jsonp'
    # Bվ��ȡ��˿����api
    while(1):
        now_time_id = time.time()
        # ��ȡ��ǰʱ���
        now_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
        # ����ʽ��ȡ��ǰʱ��
        page = urllib.request.urlopen(url)
        text = page.read()
        # ��ȡ��ҳ����
        text = text.decode('utf-8')
        i = text.find('follower')
        i = i + 10
        for j in range(i, len(text) - 1):
            if(text[j] > '9' or text[j] < '0'):
                break
        # ȡ�÷�˿��
        print(now_time)
        print('��ǰ��˿��:' + text[i: j])
        now_row = [now_time, int(text[i: j])]
        table.append(now_row)
        # ���ζ�ȡ������ĩβ
        File.save(table_name)
        while(1):
            if(time.time() - now_time_id >= stoptime):
                break
        # ��֤���δ��������ض�ʱ��

get_num_of_fans()